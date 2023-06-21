from flask import Flask, jsonify, request,render_template, send_file,redirect,url_for
import pandas as pd
import json
import geopandas as gpd
from shapely.geometry import Point
import pyproj
from geopy.distance import geodesic
import geocoder
import openpyxl
import geocoder
from geopy.distance import geodesic

region_percent = []
typeListString = ("beauty_salon","cafe","restaurants","dentist","car_repair")

def getRegion(location): #得到輸入地址為甚麼區
    for i in range(len(location)):
        if(location[i] == "區"):
            region = str(location[i-2]) + str(location[i-1]) + str(location[i])
    return region


def lonlat_origin(location): #輸出經緯度
    geocode_result = geocoder.arcgis(location)
    return geocode_result 

def transformer (location): #轉換投影後的經緯度
    # 進行地理編碼
    geocode_result = geocoder.arcgis(location)

    # Define the source and target coordinate reference systems (CRS)
    src_crs = pyproj.CRS('EPSG:4326')  # Input CRS is WGS 84
    dst_crs = pyproj.CRS('EPSG:3826')  # Target CRS is WGS 84 Web Mercator (Spherical Mercator)

    # Create a transformer object for the coordinate transformation
    transformer = pyproj.Transformer.from_crs(src_crs, dst_crs, always_xy=True)

    # Example coordinates in longitude and latitude
    lon = geocode_result.lng
    lat = geocode_result.lat

    # Transform the coordinates to the target CRS (WGS 84 Web Mercator)
    converted_coords = transformer.transform(lon, lat)

    # Print the transformed coordinates
    converted_lon = converted_coords[0]
    converted_lat = converted_coords[1]

    return converted_lat,converted_lon


def get_admin_districts_within_range (taipei_boundary, latitude, longitude, radius_km): #計算包含哪些行政區
    target_point = Point(longitude, latitude)
    target_buffer = target_point.buffer(radius_km * 1000)  # Convert to meters

    gdf_target = gpd.GeoDataFrame(geometry=[target_buffer], crs=taipei_boundary.crs)

    # Perform spatial intersection to get the administrative districts within the range
    intersected_areas = gpd.overlay(taipei_boundary, gdf_target, how='intersection')
    

    # Calculate the intersection area within each administrative district
    intersected_areas['Intersection_Area'] = intersected_areas.geometry.area

    return intersected_areas


def calculate_area(intersected_areas):  #計算包含的行政區各佔多少%及面積
    # Calculate the total area of intersected administrative districts
    total_intersected_area = intersected_areas['Intersection_Area'].sum()

    # Calculate the total area of the target range
    target_area = intersected_areas.geometry.unary_union.area

    # Print the administrative districts, their areas within the range, their proportions, and the target range area
    for index, row in intersected_areas.iterrows():
        admin_district = row['PTNAME']
        area_within_range = row['Intersection_Area']
        proportion = (area_within_range / row['AREA']) * 100

        admin_district = getRegion(admin_district)
        region_percent.append([admin_district,proportion])

    print("包含的行政區各佔多少%", region_percent)


def store_count(geocode_result,region,typeListString): #計算4km內同類型有多少店家
    store_count = []
    #位址可以換成你的檔案位址
    cafeData = openpyxl.load_workbook('./store_data/cafe.xlsx', data_only=True)
    beauty_salonData = openpyxl.load_workbook('./store_data/beauty_salon.xlsx', data_only=True)
    car_repairData = openpyxl.load_workbook('./store_data/car_repair.xlsx', data_only=True)
    dentistData = openpyxl.load_workbook('./store_data/dentist.xlsx', data_only=True)
    restaurantsData = openpyxl.load_workbook('./store_data/restaurants.xlsx', data_only=True)
    storeData = openpyxl.load_workbook('./taipei_data/taipei_cost.xlsx', data_only=True)

    cafeData = cafeData['Sheet1']
    beauty_salonData = beauty_salonData['Sheet1']
    car_repairData = car_repairData['Sheet1']
    dentistData = dentistData['Sheet1']
    restaurantsData = restaurantsData['Sheet1']
    storeData = storeData['開店上限數量']

    typeList = (beauty_salonData,cafeData,restaurantsData,dentistData,car_repairData)

    for i in range(5):
        s1 = typeList[i]
        storecount = 0 #計算4km裡同類型店家數量
        count = 2 

        while s1.cell(count,1).value != None: #如果excel表的值為空代表是最後一筆檔案了
            count +=1
            dist = geodesic((geocode_result.lat,geocode_result.lng),(s1.cell(count,4).value,s1.cell(count,3).value)).kilometers

            if dist<=4:
                storecount+=1

        store_count.append([typeListString[i],storecount])

    return store_count

def count_region_type_max(region_percent,typeListString): #根據行政區比例計算每個類型上限店家數
    type_max = []
    storeData = openpyxl.load_workbook('./taipei_data/taipei_cost.xlsx', data_only=True)
    storeData = storeData['開店上限數量']

    for j in range(2,7):
        count = 0
        for k in range(2,14):
            for l in range(len(region_percent)):
                if storeData.cell(k,1).value == region_percent[l][0]:
                    count += (storeData.cell(k,j).value)*(region_percent[l][1]/100)

        type_max.append([typeListString[j-2] , count])

    return type_max

def final_cal(store_count_re, count_region_type_max_re):
    final_re = []
    for i in range(5):
        final_re.append([store_count_re[i][0],int(count_region_type_max_re[i][1]-store_count_re[i][1])])
    # 依照數字大小做排序
    final_re_sorted = sorted(final_re, key=lambda x: x[1], reverse=True)
    return final_re_sorted


def process_location_data(location):

    location_data = location  #給使用者輸入

    region = getRegion(location_data)  #擷取使用者地址在甚麼區

    lonlat_origin_re = lonlat_origin(location_data) #計算出經緯度

    lonlat = transformer(location_data) #轉換經緯度

    # Load Taipei administrative district boundary data
    taipei_shapefile = './taipei_data/G97_A_CADIST_P.shp'
    taipei_boundary = gpd.read_file(taipei_shapefile, encoding='utf-8')
    taipei_boundary.crs = "epsg:4326"

    intersected_areas = get_admin_districts_within_range (taipei_boundary, lonlat[0], lonlat[1], 4)

    calculate_area(intersected_areas)
    store_count_re = store_count(lonlat_origin_re,region,typeListString)
    count_region_type_max_re = count_region_type_max(region_percent,typeListString)
    result_cal = final_cal(store_count_re, count_region_type_max_re)
    result_dict = {'location':result_cal}

    return result_dict


app = Flask(__name__)

status = {"status" : "Yes"}

@app.route('/status', methods=['GET'])
def get_books():

    return jsonify(status)

@app.route('/your_location', methods=['GET'])
def process_input():
    
    location = request.args.get('data')
    result = process_location_data(location)
    data = result["location"]
    
    # 讀取 Excel 檔案
    df = pd.read_excel('./taipei_data/Customer_order.xlsx')

    

    response =""
    for i in range(5):
        # 取得符合條件的資料列
        matching_rows = df[df['店家類別'].isin([data[i][0]])]
        # 取得相符合的姓名和電話欄位
        result_phone = matching_rows['電話'].values
        result_name = matching_rows['姓名'].values 
        response += str(i+1) + "."+ data[i][0] + "                                        "
        for m in range(7): 
            response += str(result_name[m]) + " "
            response += str(result_phone[m]) + "                   "

        
        
    
    response = response.replace("cafe","咖啡廳")
    response = response.replace("car_repair","汽車修理業")
    response = response.replace("restaurants","餐廳")
    response = response.replace("dentist","診所")
    response = response.replace("beauty_salon","美髮廳")

    result_dict = {'location':response}

    return result_dict


if __name__ == '__main__':
    app.run()
